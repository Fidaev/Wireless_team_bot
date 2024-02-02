from aiogram.types import InputFile
from Fidaev.Create_bot import dp, bot
import datetime
import os
import openpyxl  # библиотеки для работы с Ecxel-таблицами
from Fidaev.Project_Iskander.DATA.defla import tash_rnc1_333, tash_rnc2_333, tash_rnc1_222em, tash_rnc2_222em
from Fidaev.Project_Iskander.DATA.massivemimoandsix import tash_rnc1_222_mm, tash_rnc2_222_mm, tash_rnc1_666, tash_rnc2_666
from Fidaev.Project_Iskander.DATA.u2100u900 import tash_rnc2_u900, tash_rnc2_u2100_u900, tash_rnc2_u2100mm_u900, tash_rnc2_u2100s6_u900
from Fidaev.Project_Iskander.DATA.bts3902e import tash_rnc1_bts3902, tash_rnc2_bts3902
from Fidaev.Project_Iskander.DATA.lamp3900 import tash_rnc1_lamp3900_2, tash_rnc2_lamp3900_2, tash_rnc1_lamp3900_22, tash_rnc2_lamp3900_22, \
    tash_rnc1_lamp3900_222, tash_rnc2_lamp3900_222, tash_rnc1_lamp3900_2222, tash_rnc2_lamp3900_2222, \
    tash_rnc1_lamp3900_22222, tash_rnc2_lamp3900_22222
from Fidaev.Project_Iskander.DATA.lamp59001 import tash_rnc1_lamp5900_3, tash_rnc2_lamp5900_3, tash_rnc1_lamp5900_33, tash_rnc2_lamp5900_33, \
    tash_rnc1_lamp5900_333, tash_rnc2_lamp5900_333
from Fidaev.Project_Iskander.DATA.lamp59002 import tash_rnc1_lamp5900_3333, tash_rnc2_lamp5900_3333, tash_rnc1_lamp5900_33333, \
    tash_rnc2_lamp5900_33333
from Fidaev.Project_Iskander.DATA.moran import tash_rnc1_mobiuz_lamp5900_111, tash_rnc1_beeline_lamp5900_111, tash_rnc1_ucell_lamp5900_111


async def rnc_configuration(chad_id, omip, umts_ip, maska, configuration):
    global file_name
    today = datetime.date.today().strftime('%Y-%m-%d')

    om_ip = f"{omip}"
    umtsip = f"{umts_ip}"
    mask = f"{maska}"
    config = f"{configuration}"

    files = os.listdir(fr"/home/server/Projects/Bot_Wireless/Fidaev/Input_files/{today}/documents")
    for fayl in files:
        file_name = fayl
    kniga = openpyxl.load_workbook(
        filename=fr'/home/server/Projects/Bot_Wireless/Fidaev/Input_files/{today}/documents/{file_name}')

    UMTS_Sheet = kniga["UMTS"]  # Создаём переменную UMTS_Sheet и через нее открываем страницу LTE

    e = UMTS_Sheet.max_row + 1  # Создаем переменную "e" и задаём ей максимальное количество строк через max_row+1

    # ******************************************************BSC*************************************************************
    RNC = []
    nodebID_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    nodebNAME_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    cellNAME_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    LocallCellID_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    LAC_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    CellID_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    pci_list = []  # Создаем список/массив куда будут записываться данные с RF-плана
    arfcn_list = []  # Создаем список/массив куда будут записываться данные с RF-плана

    for col in range(0, UMTS_Sheet.max_column):
        title = UMTS_Sheet[1][col].value
        if title == "RNCNAME" or title == "RNC NAME" or title == "BSC Name" or title == "BSCName" or title == "RNC NAME " or title == "BSC Name ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                RNC.append(UMTS_Sheet[row][col].value)
        elif title == "Site ID" or title == "NodeB ID" or title == "NodeBID" or title == "SiteID" or title == "NodeB_ID" or title == "Site_ID" or title == "Site ID " or title == "NodeB ID ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                nodebID_list.append(UMTS_Sheet[row][col].value)
        elif title == "NodeB Name" or title == "Site Name(*)" or title == "Site Name( * )" or title == "Site Name(* )" or title == "Site Name( *)" or title == "Site Name()" or title == "NodeB Name " or title == "Site Name(*) ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                nodebNAME_list.append(UMTS_Sheet[row][col].value)
        elif title == "Cell Name(*)" or title == "Cell Name" or title == "CellName" or title == "Cell Name " or title == "CellName ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                cellNAME_list.append(UMTS_Sheet[row][col].value)
        elif title == "Local Cell ID" or title == "Local CellID" or title == "LocalCellID" or title == "Local Cell ID ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                LocallCellID_list.append(UMTS_Sheet[row][col].value)
        elif title == "LAC" or title == "Location Area Code" or title == "Location Area Code " or title == "LAC ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                LAC_list.append(UMTS_Sheet[row][col].value)
        elif title == "CI(*)" or title == "CI(*) " or title == "CI" or title == "Cell ID" or title == "Cell ID " or title == "CI ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                CellID_list.append(UMTS_Sheet[row][col].value)
        elif title == "ARFCN(*)" or title == "ARFCN(*) " or title == "ARFCN" or title == "ARFCN " or title == "Downlink UARFCN" or title == "Downlink UARFCN " or title == "UARFCN" or title == "UARFCN ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                arfcn_list.append(UMTS_Sheet[row][col].value)
        elif title == "PSC" or title == "DL Primary Scrambling Code" or title == "DL Primary Scrambling Code " or title == "Primary Scrambling Code" or title == "Primary Scrambling Code " or title == "PSC ":
            for row in range(2, UMTS_Sheet.max_row + 1):
                pci_list.append(UMTS_Sheet[row][col].value)

    if len(LocallCellID_list) == 0:
        for i in CellID_list:
            LocallCellID_list.append(f"{i}")

    # ************************************************************************************************************************

    if RNC[0] == "Tash_RNC1_H":  # если RNC = UMTS_Sheet['D2'].value == Tash_RNC1_H
        if config == "1":  # UMTS2100 3/3/3 Tash_RNC1
            print("ok")
            tash_rnc1_333(om_ip, umtsip, mask)

        elif config == "4":  # UMTS2100 2/2/2   EasyMacro 2.0
            tash_rnc1_222em(om_ip, umtsip, mask)

        elif config == "5":  # UMTS2100 2/2/2   MassiveMIMO
            tash_rnc1_222_mm(om_ip, umtsip, mask)

        elif config == "7":  # UMTS2100 6/6/6   6-sectors
            tash_rnc1_666(om_ip, umtsip, mask)

        elif config == "9":  # UMTS2100 2//   bts3902e
            tash_rnc1_bts3902(om_ip, umtsip, mask)

        elif config == "10":  # UMTS2100 2//   LampSite3900
            tash_rnc1_lamp3900_2(om_ip, umtsip, mask)

        elif config == "11":  # UMTS2100 2/2/   LampSite3900
            tash_rnc1_lamp3900_22(om_ip, umtsip, mask)

        elif config == "12":  # UMTS2100 2/2/2   LampSite3900
            tash_rnc1_lamp3900_222(om_ip, umtsip, mask)

        elif config == "13":  # UMTS2100 2/2/2/2   LampSite3900
            tash_rnc1_lamp3900_2222(om_ip, umtsip, mask)

        elif config == "14":  # UMTS2100 2/2/2/2/2  LampSite3900
            tash_rnc1_lamp3900_22222(om_ip, umtsip, mask)

        elif config == "15":  # UMTS2100 3//  LampSite5900
            tash_rnc1_lamp5900_3(om_ip, umtsip, mask)

        elif config == "16":  # UMTS2100 3/3/  LampSite5900
            tash_rnc1_lamp5900_33(om_ip, umtsip, mask)

        elif config == "17":  # UMTS2100 3/3/3  LampSite5900
            tash_rnc1_lamp5900_333(om_ip, umtsip, mask)

        elif config == "18":  # UMTS2100 3/3/3/3  LampSite5900
            tash_rnc1_lamp5900_3333(om_ip, umtsip, mask)

        elif config == "19":  # UMTS2100 3/3/3/3/3  LampSite5900
            tash_rnc1_lamp5900_33333(om_ip, umtsip, mask)

        elif config == "20":  # UMTS2100 1/1/1  LampSite5900 MobiUZ
            tash_rnc1_mobiuz_lamp5900_111(om_ip, umtsip, mask)

        elif config == "21":  # UMTS2100 1/1/1  LampSite5900 Beeline
            tash_rnc1_beeline_lamp5900_111(om_ip, umtsip, mask)

        elif config == "22":  # UMTS2100 1/1/1  LampSite5900 Ucell
            tash_rnc1_ucell_lamp5900_111(om_ip, umtsip, mask)

    elif RNC[0] == "Tash_RNC2_H":  # если RNC = UMTS_Sheet['D2'].value == Tash_RNC2_H
        if config == "1":  # UMTS2100 3/3/3 Tash_RNC2
            tash_rnc2_333(om_ip, umtsip, mask)

        elif config == "4":  # UMTS2100 2/2/2   EasyMacro 2.0 Tash_RNC2
            tash_rnc2_222em(om_ip, umtsip, mask)

        elif config == "5":  # UMTS2100 2/2/2   MassiveMIMO Tash_RNC2
            tash_rnc2_222_mm(om_ip, umtsip, mask)

        elif config == "7":  # UMTS2100 6/6/6   6-sectors
            tash_rnc2_666(om_ip, umtsip, mask)

        elif config == "2":  # UMTS900 1/1/1
            tash_rnc2_u900(om_ip, umtsip, mask)

        elif config == "3":  # UMTS2100 3/3/3 UMTS900 1/1/1
            tash_rnc2_u2100_u900(om_ip, umtsip, mask)

        elif config == "6":  # UMTS2100 3/3/3 UMTS900 1/1/1
            tash_rnc2_u2100mm_u900(om_ip, umtsip, mask)

        elif config == "8":  # UMTS2100 6/6/6 UMTS900 1/1/1
            tash_rnc2_u2100s6_u900(om_ip, umtsip, mask)

        elif config == "9":  # UMTS2100 2//   bts3902e
            tash_rnc2_bts3902(om_ip, umtsip, mask)

        elif config == "10":  # UMTS2100 2//   LampSite3900
            tash_rnc2_lamp3900_2(om_ip, umtsip, mask)

        elif config == "11":  # UMTS2100 2/2/   LampSite3900
            tash_rnc2_lamp3900_22(om_ip, umtsip, mask)

        elif config == "12":  # UMTS2100 2/2/2   LampSite3900
            tash_rnc2_lamp3900_222(om_ip, umtsip, mask)

        elif config == "13":  # UMTS2100 2/2/2/2   LampSite3900
            tash_rnc2_lamp3900_2222(om_ip, umtsip, mask)

        elif config == "14":  # UMTS2100 2/2/2/2/2  LampSite3900
            tash_rnc2_lamp3900_22222(om_ip, umtsip, mask)

        elif config == "15":  # UMTS2100 3//  LampSite5900
            tash_rnc2_lamp5900_3(om_ip, umtsip, mask)

        elif config == "16":  # UMTS2100 3/3/  LampSite5900
            tash_rnc2_lamp5900_33(om_ip, umtsip, mask)

        elif config == "17":  # UMTS2100 3/3/3  LampSite5900
            tash_rnc2_lamp5900_333(om_ip, umtsip, mask)

        elif config == "18":  # UMTS2100 3/3/3/3  LampSite5900
            tash_rnc2_lamp5900_3333(om_ip, umtsip, mask)

        elif config == "19":  # UMTS2100 3/3/3/3/3  LampSite5900
            tash_rnc2_lamp5900_33333(om_ip, umtsip, mask)

    files = os.listdir(fr"/home/server/Projects/Bot_Wireless/Fidaev/Output_files")
    for fayl in files:
        file_name = fayl
    sending_file = InputFile(fr"/home/server/Projects/Bot_Wireless/Fidaev/Output_files/{file_name}")
    await bot.send_document(chat_id=chad_id, document=sending_file)
